# copied from here
# https://docs.gurobi.com/projects/examples/en/current/examples/diet.html#diet4
# maintaining a copy here just in case the Gurobi example goes down
# note that this file assumes a cross tab representation of the
# Nutrition Quantities table. ticdat uses normalized table data in Excel (i.e. the first
# row has column names, not actual data)
#!/usr/bin/env python3.11

# Copyright 2025, Gurobi Optimization, LLC

# Read diet model data from an Excel spreadsheet (diet.xlsx).
# Pass the imported data into the diet model (dietmodel.py).
#
# Note that this example reads an external data file (..\data\diet.xlsx).
# As a result, it must be run from the Gurobi examples/python directory.
#
# This example uses Python package 'openpyxl', which isn't included
# in most Python distributions.  You can install it with
# 'pip install openpyxl'.

import os
import openpyxl
import dietmodel

# Open 'diet.xlsx'
book = openpyxl.load_workbook(os.path.join("..", "data", "diet.xlsx"))

# Read min/max nutrition info from 'Categories' sheet
sheet = book["Categories"]
categories = []
minNutrition = {}
maxNutrition = {}
for row in sheet.iter_rows():
    category = row[0].value
    if category != "Categories":
        categories.append(category)
        minNutrition[category] = row[1].value
        maxNutrition[category] = row[2].value

# Read food costs from 'Foods' sheet
sheet = book["Foods"]
foods = []
cost = {}
for row in sheet.iter_rows():
    food = row[0].value
    if food != "Foods":
        foods.append(food)
        cost[food] = row[1].value

# Read food nutrition info from 'Nutrition' sheet
sheet = book["Nutrition"]
nutritionValues = {}
for row in sheet.iter_rows():
    if row[0].value == None:  # column labels - categories
        cats = [v.value for v in row]
    else:  # nutrition values
        food = row[0].value
        for col in range(1, len(row)):
            nutritionValues[food, cats[col]] = row[col].value


dietmodel.solve(categories, minNutrition, maxNutrition, foods, cost, nutritionValues)
